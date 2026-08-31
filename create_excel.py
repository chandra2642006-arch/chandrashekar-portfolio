import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

data = [
    (1, "Srinath", 212),
    (2, "Mohan", 243),
    (3, "Vishnu Karthik", 241),
    (4, "Chandrasekhar", 234),
    (5, "Vinay", 242),
    (6, "Lakshya Hassini", 232),
    (7, "Venkatesh", 265),
    (8, "Charan", 251),
    (9, "Vardhan Arjun", 210),
    (10, "Venkat", 279),
    (11, "Sri Charan", 217),
    (12, "Tejaswini", 244)
]

# 1. Save CSV
csv_file = "Student_Team_Details.csv"
with open(csv_file, mode="w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(["S.No.", "Student Name", "Registration Number"])
    for row in data:
        writer.writerow(row)

print(f"CSV file created: {csv_file}")

# 2. Create XLSX with OpenPyXL
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Team Members"

# Enable grid lines explicitly
ws.views.sheetView[0].showGridLines = True

# Colors
NAVY_HEADER = "1F4E78"
WHITE = "FFFFFF"
LIGHT_ZEBRA = "F2F5F9"
BORDER_COLOR = "D9D9D9"
TITLE_COLOR = "0D233A"

# Fonts
title_font = Font(name="Calibri", size=16, bold=True, color=TITLE_COLOR)
subtitle_font = Font(name="Calibri", size=10, italic=True, color="595959")
header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
data_font = Font(name="Calibri", size=11)
bold_font = Font(name="Calibri", size=11, bold=True)

# Fills
header_fill = PatternFill(start_color=NAVY_HEADER, end_color=NAVY_HEADER, fill_type="solid")
zebra_fill = PatternFill(start_color=LIGHT_ZEBRA, end_color=LIGHT_ZEBRA, fill_type="solid")

# Alignments
align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")

# Borders
thin_border = Border(
    left=Side(style='thin', color=BORDER_COLOR),
    right=Side(style='thin', color=BORDER_COLOR),
    top=Side(style='thin', color=BORDER_COLOR),
    bottom=Side(style='thin', color=BORDER_COLOR)
)
header_border = Border(
    left=Side(style='thin', color="FFFFFF"),
    right=Side(style='thin', color="FFFFFF"),
    top=Side(style='medium', color=NAVY_HEADER),
    bottom=Side(style='medium', color=NAVY_HEADER)
)
total_border = Border(
    top=Side(style='thin', color="000000"),
    bottom=Side(style='double', color="000000")
)

# Title Block
ws.merge_cells("A1:C1")
ws["A1"] = "Student Team Roster"
ws["A1"].font = title_font
ws["A1"].alignment = align_left

ws.merge_cells("A2:C2")
ws["A2"] = "Detailed List of Team Members & Registration Numbers"
ws["A2"].font = subtitle_font
ws["A2"].alignment = align_left

ws.row_dimensions[1].height = 25
ws.row_dimensions[2].height = 18
ws.row_dimensions[3].height = 10  # blank row space

# Table Header (Row 4)
headers = ["S.No.", "Student Name", "Registration Number"]
ws.row_dimensions[4].height = 24

for col_num, header_title in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_num, value=header_title)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = align_center
    cell.border = header_border

# Data Rows (Row 5 to 16)
for row_idx, row_data in enumerate(data, start=5):
    ws.row_dimensions[row_idx].height = 20
    s_no, name, reg_no = row_data
    
    c1 = ws.cell(row=row_idx, column=1, value=s_no)
    c1.alignment = align_center
    
    c2 = ws.cell(row=row_idx, column=2, value=name)
    c2.alignment = align_left
    
    c3 = ws.cell(row=row_idx, column=3, value=reg_no)
    c3.alignment = align_center
    
    # Apply fonts, borders, zebra striping
    for col_num in range(1, 4):
        cell = ws.cell(row=row_idx, column=col_num)
        cell.font = data_font
        cell.border = thin_border
        if (row_idx % 2) == 0:
            cell.fill = zebra_fill

# Summary Row (Row 17)
summary_row = len(data) + 5
ws.row_dimensions[summary_row].height = 22

ws.cell(row=summary_row, column=1, value="").border = total_border

lbl_cell = ws.cell(row=summary_row, column=2, value="Total Team Members")
lbl_cell.font = bold_font
lbl_cell.alignment = align_left
lbl_cell.border = total_border

val_cell = ws.cell(row=summary_row, column=3, value=f"=COUNTA(C5:C{summary_row-1})")
val_cell.font = bold_font
val_cell.alignment = align_center
val_cell.border = total_border

# Auto-adjust column widths
for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        # Ignore merged title row length for column width calc
        if cell.row in [1, 2]:
            continue
        val_str = str(cell.value or "")
        if len(val_str) > max_len:
            max_len = len(val_str)
    ws.column_dimensions[col_letter].width = max(max_len + 5, 14)

xlsx_file = "Student_Team_Details.xlsx"
wb.save(xlsx_file)
print(f"Excel file created: {xlsx_file}")
